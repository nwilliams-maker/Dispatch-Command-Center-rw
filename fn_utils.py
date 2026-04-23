"""
fn_utils.py — Terraboost Media Field Nation Utilities
All Field Nation logic lives here: manager mapping, upload generation, background saves.
"""

import io
import threading
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ---------------------------------------------------------------------------
# State → Work Order Manager (by pod)
# ---------------------------------------------------------------------------
FN_STATE_MANAGER = {
    # Orange Pod
    "AK": "Bernice Makaya", "AZ": "Bernice Makaya", "CA": "Bernice Makaya",
    "HI": "Bernice Makaya", "ID": "Bernice Makaya", "NV": "Bernice Makaya",
    "OR": "Bernice Makaya", "WA": "Bernice Makaya",
    # Green Pod
    "CO": "Reabetswe Segopa", "DC": "Reabetswe Segopa", "GA": "Reabetswe Segopa",
    "IN": "Reabetswe Segopa", "KY": "Reabetswe Segopa", "MD": "Reabetswe Segopa",
    "NJ": "Reabetswe Segopa", "OH": "Reabetswe Segopa", "UT": "Reabetswe Segopa",
    # Red Pod
    "CT": "Lee Adams", "DE": "Lee Adams", "MA": "Lee Adams", "ME": "Lee Adams",
    "NH": "Lee Adams", "NY": "Lee Adams", "PA": "Lee Adams", "RI": "Lee Adams",
    "VA": "Lee Adams", "VT": "Lee Adams", "WV": "Lee Adams",
    # Blue Pod
    "AL": "Elna Burger", "AR": "Elna Burger", "FL": "Elna Burger", "IA": "Elna Burger",
    "IL": "Elna Burger", "LA": "Elna Burger", "MI": "Elna Burger", "MN": "Elna Burger",
    "MO": "Elna Burger", "MS": "Elna Burger", "NC": "Elna Burger", "SC": "Elna Burger",
    "WI": "Elna Burger",
    # Purple Pod
    "KS": "Stacey Ferreira", "MT": "Stacey Ferreira", "ND": "Stacey Ferreira",
    "NE": "Stacey Ferreira", "NM": "Stacey Ferreira", "OK": "Stacey Ferreira",
    "SD": "Stacey Ferreira", "TN": "Stacey Ferreira", "TX": "Stacey Ferreira",
    "WY": "Stacey Ferreira",
}

PAY_PER_STOP = 20.0


# ---------------------------------------------------------------------------
# Background sheet save — never blocks the UI
# ---------------------------------------------------------------------------
def save_fn_to_sheet(gas_url: str, payload: dict, session_state=None) -> None:
    """Fire-and-forget: saves a route to the Field Nation Google Sheet tab.
    Clears the reverted flag from session_state after the write completes."""
    cluster_hash = payload.get("cluster_hash")

    def _worker():
        try:
            requests.post(gas_url, json={"action": "saveToFieldNation", "payload": payload}, timeout=15)
        except Exception:
            pass
        finally:
            # Clear reverted flag once sheet write is done (success or fail)
            if session_state is not None and cluster_hash:
                session_state.pop(f"reverted_{cluster_hash}", None)

    threading.Thread(target=_worker, daemon=True).start()


# ---------------------------------------------------------------------------
# Mass upload file generator
# ---------------------------------------------------------------------------
def generate_fn_upload(stop_metrics: dict, cluster: dict, due, final_pay: float, cluster_hash: str):
    """
    Generates a Field Nation mass upload Excel file.

    Structure:
      - One row per stop address
      - One numbered slot (1–5) per unique locationinVenue at that stop
      - $20 fixed per stop (PAY_PER_STOP)
      - Work Order Manager resolved by state via FN_STATE_MANAGER

    Custom field mapping per slot:
      N. Customer Name   → clientCompany
      1. Venue ID        → venueId  (slot 1 only)
      N. Location in Venue → taskType + " — " + locationinVenue

    Returns:
      (BytesIO buffer, int stop_count)  or  (None, 0) if no kiosk stops found.
    """

    # Build address → [tasks] grouped by unique locationinVenue
    stop_kiosk_tasks: dict = {}
    for t in cluster.get('data', []):
        tt = str(t.get('task_type', '')).lower()
        if 'install' not in tt and 'kiosk' not in tt:
            continue
        addr = t.get('full', '')
        if addr not in stop_kiosk_tasks:
            stop_kiosk_tasks[addr] = []
        loc = t.get('location_in_venue', '').strip()
        existing = [x.get('location_in_venue', '') for x in stop_kiosk_tasks[addr]]
        if loc not in existing:
            stop_kiosk_tasks[addr].append(t)

    kiosk_stops = [(addr, tasks) for addr, tasks in stop_kiosk_tasks.items() if tasks]
    if not kiosk_stops:
        return None, 0

    # Format due date
    try:
        if hasattr(due, 'strftime'):
            start_date = due.strftime("%-m/%-d/%Y")
            end_date   = due.strftime("%-m/%-d/%Y")
        else:
            due_dt     = datetime.strptime(str(due), "%Y-%m-%d")
            start_date = due_dt.strftime("%-m/%-d/%Y")
            end_date   = due_dt.strftime("%-m/%-d/%Y")
    except Exception:
        start_date = str(due)
        end_date   = str(due)

    # Headers: base columns + 5 numbered custom field sets
    base_headers = [
        "Location Name", "Address #1", "City", "State", "Postal Code", "Country",
        "Schedule Type", "Scheduled Start Date", "Scheduled Start Time",
        "Scheduled End Date", "Scheduled End Time", "Pay Type", "Pay Rate",
        "Approximate Hours to Complete", "Est. WO-Value", "Work Order Manager", "",
    ]
    custom_headers = []
    for n in range(1, 6):
        custom_headers.append(f"{n}. Customer Name")
        if n == 1:
            custom_headers.append("1. Venue ID")
        custom_headers.append(f"{n}. Location in Venue")

    all_headers = base_headers + custom_headers

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_fill = PatternFill("solid", start_color="FEF9C3")
    header_font = Font(bold=True, name="Arial")
    for col, h in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (addr, tasks) in enumerate(kiosk_stops, 2):
        parts    = [p.strip() for p in addr.split(",")]
        street   = parts[0] if len(parts) > 0 else addr
        city     = parts[1] if len(parts) > 1 else cluster.get('city', '')
        state    = parts[2].strip().upper() if len(parts) > 2 else cluster.get('state', '')
        zip_code = tasks[0].get('zip', parts[3].strip() if len(parts) > 3 else '')

        venue_name = tasks[0].get('venue_name', 'Terraboost Media')
        manager    = FN_STATE_MANAGER.get(state, '')

        base_row = [
            venue_name,
            street,
            city,
            state,
            zip_code,
            "US",
            "Complete work anytime over a date range",
            start_date,
            "8:00 AM",
            end_date,
            "5:00 PM",
            "Fixed",
            PAY_PER_STOP,
            1.0,
            PAY_PER_STOP,
            manager,
            "",
        ]

        custom_cols = []
        for slot_idx, task in enumerate(tasks[:5], 1):
            task_type    = str(task.get('task_type', 'Kiosk Install')).strip()
            loc_in_venue = str(task.get('location_in_venue', '')).strip()
            client       = str(task.get('client_company', 'Terraboost Media')).strip()
            venue_id     = str(task.get('venue_id', '')).strip()
            combined_loc = f"{task_type} — {loc_in_venue}" if loc_in_venue else task_type

            custom_cols.append(client)
            if slot_idx == 1:
                custom_cols.append(venue_id)
            custom_cols.append(combined_loc)

        # Pad empty slots up to 5
        filled = len(tasks[:5])
        for slot_idx in range(filled + 1, 6):
            custom_cols.append("")
            if slot_idx == 1:
                custom_cols.append("")
            custom_cols.append("")

        full_row = base_row + custom_cols
        for col, val in enumerate(full_row, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = Font(name="Arial")

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, len(kiosk_stops)
